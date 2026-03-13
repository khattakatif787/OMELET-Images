import os
import argparse
import numpy as np
import pandas as pd
from openpyxl import load_workbook

REJECT_TAG = "-1"

SCC_FOLDER = "debug/SCC_outputs"
OUT_FOLDER = "debug/ENSEMBLE_outputs"

MASTER_XLSX = os.path.join(OUT_FOLDER, "master_results.xlsx")
VAL_CSV = os.path.join(OUT_FOLDER, "master_results_Validation.csv")
TEST_CSV = os.path.join(OUT_FOLDER, "master_results_Test.csv")


def compute_metrics(true_labels: np.ndarray, preds: np.ndarray) -> dict:

    accepted = preds != REJECT_TAG
    correct = np.sum(accepted & (preds == true_labels))
    misclass = np.sum(accepted & (preds != true_labels))
    rejected = np.sum(~accepted)

    total = len(true_labels)

    acc = correct / total if total > 0 else 0.0

    return {
        "correct": int(correct),
        "misclass": int(misclass),
        "rejected": int(rejected),
        "total": int(total),
        "ensemble_accuracy": float(acc),
    }


def pick_first_fcc_by_aw(true_labels: np.ndarray, fcc_preds: pd.DataFrame) -> str:

    best_name = None
    best_aw = -1.0

    for col in fcc_preds.columns:

        preds = fcc_preds[col].astype(str).to_numpy()

        aw = np.sum(preds == true_labels) / len(true_labels)

        if aw > best_aw:

            best_aw = aw
            best_name = col

    return best_name


def build_ensemble_gain_only(true_labels: np.ndarray, fcc_preds: pd.DataFrame, max_fccs: int) -> list[str]:

    cols = list(fcc_preds.columns)

    if len(cols) == 0 or max_fccs <= 0:
        return []

    first = pick_first_fcc_by_aw(true_labels, fcc_preds)

    ensemble_order = [first]

    remaining = [c for c in cols if c != first]

    ensemble_preds = np.full_like(true_labels, fill_value=REJECT_TAG, dtype=object)

    first_preds = fcc_preds[first].astype(str).to_numpy()

    mask_accept = first_preds != REJECT_TAG

    ensemble_preds[mask_accept] = first_preds[mask_accept]

    while remaining and len(ensemble_order) < max_fccs:

        gains = {}

        for c in remaining:

            p = fcc_preds[c].astype(str).to_numpy()

            gain = np.sum((ensemble_preds == REJECT_TAG) & (p == true_labels))

            gains[c] = int(gain)

        candidates = [c for c, g in gains.items() if g > 0]

        if not candidates:
            break

        best = max(candidates, key=lambda c: gains[c])

        ensemble_order.append(best)

        remaining.remove(best)

        pbest = fcc_preds[best].astype(str).to_numpy()

        update_mask = (ensemble_preds == REJECT_TAG) & (pbest != REJECT_TAG)

        ensemble_preds[update_mask] = pbest[update_mask]

    return ensemble_order


def apply_ensemble_order(df: pd.DataFrame, ensemble_order: list[str]) -> np.ndarray:

    true_labels = df.iloc[:, 0].astype(str).to_numpy()

    if not ensemble_order:
        return np.full_like(true_labels, fill_value=REJECT_TAG, dtype=object)

    mat = df[ensemble_order].astype(str).to_numpy()

    out = np.full_like(true_labels, fill_value=REJECT_TAG, dtype=object)

    for i in range(mat.shape[0]):

        for j in range(mat.shape[1]):

            if mat[i, j] != REJECT_TAG:

                out[i] = mat[i, j]

                break

    return out


def append_to_master(sheet_name, df_to_append, master_path):

    os.makedirs(os.path.dirname(master_path), exist_ok=True)

    if os.path.exists(master_path):

        with pd.ExcelWriter(master_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:

            book = writer.book

            if sheet_name in book.sheetnames:

                startrow = book[sheet_name].max_row

                df_to_append.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    startrow=startrow,
                    header=False,
                    index=False
                )

            else:

                df_to_append.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False
                )

    else:

        with pd.ExcelWriter(master_path, engine="openpyxl", mode="w") as writer:

            df_to_append.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )


def load_scc_files(fcc_alr: float):

    alr_tag = str(fcc_alr)

    val_path = os.path.join(SCC_FOLDER, f"SCC_VALIDATION_ALR_{alr_tag}.csv")
    test_path = os.path.join(SCC_FOLDER, f"SCC_TEST_ALR_{alr_tag}.csv")

    if not os.path.exists(val_path):
        raise FileNotFoundError(f"Missing: {val_path}")

    if not os.path.exists(test_path):
        raise FileNotFoundError(f"Missing: {test_path}")

    df_val = pd.read_csv(val_path)
    df_test = pd.read_csv(test_path)

    return df_val, df_test


def run_for_pair(dataset: str, fcc_alr: float, system_alr: float):

    df_val, df_test = load_scc_files(fcc_alr)

    true_val = df_val.iloc[:, 0].astype(str).to_numpy()

    fcc_preds_val = df_val.iloc[:, 1:]

    max_fccs = int(np.floor(system_alr / fcc_alr))

    if max_fccs < 1:
        max_fccs = 1

    ensemble_order = build_ensemble_gain_only(true_val, fcc_preds_val, max_fccs=max_fccs)

    first = ensemble_order[0] if ensemble_order else None

    val_ensemble_preds = apply_ensemble_order(df_val, ensemble_order)

    val_metrics_ens = compute_metrics(true_val, val_ensemble_preds)

    if first is not None:

        first_preds_val = df_val[first].astype(str).to_numpy()

        val_metrics_first = compute_metrics(true_val, first_preds_val)

    else:

        val_metrics_first = {"ensemble_accuracy": 0.0, "correct": 0, "misclass": 0}

    val_row = {
        "dataset": dataset,
        "system_alr": system_alr,
        "fcc_alr": fcc_alr,
        "num_fccs in ensemble": len(ensemble_order),
        "ensemble_order": ", ".join(ensemble_order),
        **val_metrics_ens,
        "first_fcc_accuracy": float(val_metrics_first["ensemble_accuracy"]),
        "first_fcc_correct": int(val_metrics_first["correct"]),
        "first_fcc_misclass": int(val_metrics_first["misclass"]),
    }

    val_df_out = pd.DataFrame([val_row])

    true_test = df_test.iloc[:, 0].astype(str).to_numpy()

    test_ensemble_preds = apply_ensemble_order(df_test, ensemble_order)

    test_metrics_ens = compute_metrics(true_test, test_ensemble_preds)

    if first is not None and first in df_test.columns:

        first_preds_test = df_test[first].astype(str).to_numpy()

        test_metrics_first = compute_metrics(true_test, first_preds_test)

    else:

        test_metrics_first = {"ensemble_accuracy": 0.0, "correct": 0, "misclass": 0}

    test_row = {
        "dataset": dataset,
        "system_alr": system_alr,
        "fcc_alr": fcc_alr,
        "num_fccs in ensemble": len(ensemble_order),
        "ensemble_order": ", ".join(ensemble_order),
        **test_metrics_ens,
        "first_fcc_accuracy": float(test_metrics_first["ensemble_accuracy"]),
        "first_fcc_correct": int(test_metrics_first["correct"]),
        "first_fcc_misclass": int(test_metrics_first["misclass"]),
    }

    test_df_out = pd.DataFrame([test_row])

    return val_df_out, test_df_out


def append_to_csv(csv_path: str, df_to_append: pd.DataFrame):

    os.makedirs(os.path.dirname(csv_path), exist_ok=True)

    write_header = not os.path.exists(csv_path)

    df_to_append.to_csv(csv_path, mode="a", header=write_header, index=False)


def main():

    parser = argparse.ArgumentParser()

    parser.add_argument("--dataset", type=str, required=True)

    parser.add_argument("--fcc_alrs", type=float, nargs="+", required=True)

    parser.add_argument("--ensemble_sizes", type=int, nargs="+", required=True)

    args = parser.parse_args()

    os.makedirs(OUT_FOLDER, exist_ok=True)

    for fcc_alr in args.fcc_alrs:

        for k in args.ensemble_sizes:

            system_alr = fcc_alr * k

            val_df, test_df = run_for_pair(args.dataset, fcc_alr, system_alr)

            append_to_master("Validation", val_df, MASTER_XLSX)
            append_to_master("Test", test_df, MASTER_XLSX)

            append_to_csv(VAL_CSV, val_df)
            append_to_csv(TEST_CSV, test_df)

            print(f"[OK] fcc_alr={fcc_alr} ensemble_size={k} system_alr={system_alr}")


if __name__ == "__main__":
    main()